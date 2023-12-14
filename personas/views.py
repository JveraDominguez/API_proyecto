from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.views import View
from django.views.generic import TemplateView
from openpyxl import Workbook

from .forms import CrearAutomovilForm, CrearPersonaForm
from .models import Persona, automovile


# Create your views here.
# views.py

def ver_persona(request, persona_id):
    persona = get_object_or_404(Persona, pk=persona_id)
    automoviles = persona.automoviles.all()
    datos = {'persona': persona, 'automoviles': automoviles}
    return render(request, 'ver_auto.html', datos)

def eliminar_persona(request, persona_id):
    persona = get_object_or_404(Persona, pk=persona_id)

    # Obtenemos todos los autos asociados a esta persona
    autos_asociados = automovile.objects.filter(propietario=persona)

    if request.method == 'POST':
        # Eliminamos la persona y todos sus autos asociados
        persona.delete()
        autos_asociados.delete()

        return redirect('inicio')  # Reemplaza 'bienvenida' con el nombre de tu URL de bienvenida

    return render(request, 'eliminar_persona.html', {'persona': persona, 'autos_asociados': autos_asociados})


def crear_persona(request):
    if request.method == 'POST':
        form = CrearPersonaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('inicio')  # Reemplaza con la URL correcta
    else:
        form = CrearPersonaForm()

    return render(request, 'crear_persona.html', {'form': form})


def crear_automovil(request):
    if request.method == 'POST':
        form = CrearAutomovilForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('inicio')  # Reemplaza con la URL correcta
    else:
        form = CrearAutomovilForm()

    return render(request, 'crear_automovil.html', {'form': form})

def editar_persona(request, persona_id):
    persona = get_object_or_404(Persona, pk=persona_id)

    if request.method == 'POST':
        form_persona = CrearPersonaForm(request.POST, instance=persona)
        if form_persona.is_valid():
            form_persona.save()
            return redirect('inicio')

    else:
        form_persona = CrearPersonaForm(instance=persona)

    return render(request, 'editar_persona.html', {'form_persona': form_persona, 'persona': persona})


def editar_automovil(request, automovil_id):
    automovil = get_object_or_404(automovile, pk=automovil_id)

    if request.method == 'POST':
        form_automovil = CrearAutomovilForm(request.POST, instance=automovil)
        if form_automovil.is_valid():
            form_automovil.save()
            return redirect('inicio')

    else:
        form_automovil = CrearAutomovilForm(instance=automovil)

    return render(request, 'editar_automovil.html', {'form_automovil': form_automovil, 'automovil': automovil})


class ReportePersonasExcel(View):
    def get(self, request, *args, **kwargs):
        # Obtenemos todas las personas de nuestra base de datos
        personas = Persona.objects.all()
        # Creamos el libro de trabajo
        wb = Workbook()
        # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        # En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws['B1'] = 'REPORTE DE PERSONAS'
        # Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('B1:E1')
        # Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = 'Cedula'
        ws['C3'] = 'NOMBRE'
        ws['D3'] = 'APELLIDO'
        cont = 4
        # Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        for persona in personas:
            ws.cell(row=cont, column=2).value = persona.cedula
            ws.cell(row=cont, column=3).value = persona.nombre
            ws.cell(row=cont, column=4).value = persona.apellido
            cont += 1
        # Establecemos el nombre del archivo
        nombre_archivo = "reporte de personas ingresadas.xlsx"
        # Definimos que el tipo de respuesta a devolver es un archivo de Microsoft Excel
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
